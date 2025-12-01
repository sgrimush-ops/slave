"""
Módulo para calcular sugestões de pedido baseado em vendas e estoque
"""
import pandas as pd
import numpy as np
from typing import Dict, Tuple
from pathlib import Path
import math


class CalculadorPedido:
    """Calcula sugestões de pedido considerando vendas, estoque e embalagem"""
    
    def __init__(self, dias_cobertura: int = 4, margem_seguranca: float = 1.2):
        """
        Inicializa o calculador
        
        Args:
            dias_cobertura: Dias de cobertura desejados (padrão: 4 dias)
            margem_seguranca: Margem de segurança adicional (padrão: 20%)
        """
        self.dias_cobertura = dias_cobertura
        self.margem_seguranca = margem_seguranca
    
    def calcular_sugestao_pedido(
        self,
        estoque_atual: int,
        venda_media_dia: float,
        embalagem: int,
        venda_7dias: int = 0,
        venda_14dias: int = 0,
        venda_30dias: int = 0,
        venda_60dias: int = 0,
        ponto_pedido: int = None,
        estoque_ideal: int = None
    ) -> Dict:
        """
        Calcula a sugestão de pedido para um produto
        
        Args:
            estoque_atual: Estoque atual em unidades
            venda_media_dia: Venda média diária
            embalagem: Quantidade de unidades por caixa
            venda_7dias: Venda acumulada dos últimos 7 dias
            venda_14dias: Venda acumulada dos últimos 14 dias
            venda_30dias: Venda acumulada dos últimos 30 dias
            venda_60dias: Venda acumulada dos últimos 60 dias
            ponto_pedido: Mínimo definido pelo comprador (gatilho de pedido)
            estoque_ideal: Máximo definido pelo comprador (exposição visual)
            
        Returns:
            Dicionário com sugestão e análise
        """
        
        # 1. ESTRATÉGIA INTELIGENTE: Equilibrar giro saudável (4-6 dias) com exposição visual
        
        # Calcula necessidade baseada em giro saudável
        dias_giro_minimo = 4  # Mínimo recomendado
        dias_giro_maximo = 6  # Máximo recomendado
        
        necessidade_giro_min = venda_media_dia * dias_giro_minimo
        necessidade_giro_max = venda_media_dia * dias_giro_maximo
        
        # 2. Considera valores do comprador (experiência visual)
        if ponto_pedido and estoque_ideal:
            # O comprador definiu valores baseados em experiência de exposição
            # Validar se esses valores são compatíveis com giro saudável
            
            diferenca_comprador = estoque_ideal - ponto_pedido
            dias_cobertura_comprador = diferenca_comprador / venda_media_dia if venda_media_dia > 0 else 0
            
            # Se os valores do comprador resultam em giro muito lento (>6 dias)
            if dias_cobertura_comprador > dias_giro_maximo:
                # Ajusta para o máximo recomendado, mas respeitando múltiplos de embalagem
                quantidade_necessaria = necessidade_giro_max * self.margem_seguranca - estoque_atual
                estrategia = 'giro_otimizado'
                observacao = f'Ajustado de {dias_cobertura_comprador:.1f} para {dias_giro_maximo} dias (giro mais saudável)'
            
            # Se os valores do comprador resultam em giro muito rápido (<4 dias)
            elif dias_cobertura_comprador < dias_giro_minimo:
                # Ajusta para o mínimo recomendado
                quantidade_necessaria = necessidade_giro_min * self.margem_seguranca - estoque_atual
                estrategia = 'giro_otimizado'
                observacao = f'Ajustado de {dias_cobertura_comprador:.1f} para {dias_giro_minimo} dias (evitar excesso de pedidos)'
            
            # Se está dentro do range (4-6 dias), respeita valores do comprador
            else:
                # Verifica se está abaixo do ponto de pedido
                if estoque_atual < ponto_pedido:
                    quantidade_necessaria = estoque_ideal - estoque_atual
                    estrategia = 'comprador'
                    observacao = f'Respeitando valores do comprador ({dias_cobertura_comprador:.1f} dias de cobertura)'
                else:
                    quantidade_necessaria = 0
                    estrategia = 'comprador'
                    observacao = 'Estoque acima do ponto de pedido definido'
        else:
            # Sem valores do comprador, usa apenas lógica de giro saudável
            quantidade_necessaria = necessidade_giro_min * self.margem_seguranca - estoque_atual
            estrategia = 'giro_saudavel'
            observacao = f'Baseado em giro de {dias_giro_minimo} dias (padrão do sistema)'
        
        # 3. Se não precisa pedir (estoque suficiente), retorna 0
        if quantidade_necessaria <= 0:
            return {
                'sugestao_unidades': 0,
                'sugestao_caixas': 0,
                'estoque_suficiente': True,
                'dias_cobertura_atual': estoque_atual / venda_media_dia if venda_media_dia > 0 else float('inf'),
                'estrategia': estrategia,
                'observacao': observacao,
                'ponto_pedido': ponto_pedido,
                'estoque_ideal': estoque_ideal,
                'motivo': 'Estoque atual suficiente para cobertura'
            }
        
        # 4. Arredonda para cima em múltiplos da embalagem
        caixas_necessarias = math.ceil(quantidade_necessaria / embalagem)
        sugestao_unidades = caixas_necessarias * embalagem
        
        # 5. Análise de tendência
        tendencia = self._analisar_tendencia(venda_7dias, venda_14dias, venda_30dias, venda_60dias)
        
        # 6. Ajusta baseado na tendência
        ajuste_tendencia = ''
        if tendencia['tipo'] == 'crescimento_forte':
            # Aumenta em 1 caixa se crescimento forte
            caixas_necessarias += 1
            sugestao_unidades = caixas_necessarias * embalagem
            ajuste_tendencia = ' + Ajustado para cima (crescimento forte)'
        elif tendencia['tipo'] == 'queda_forte':
            # Diminui em 1 caixa se queda forte (mas não menos que 1)
            if caixas_necessarias > 1:
                caixas_necessarias -= 1
                sugestao_unidades = caixas_necessarias * embalagem
                ajuste_tendencia = ' + Ajustado para baixo (queda forte)'
        
        motivo = observacao + ajuste_tendencia
        
        return {
            'sugestao_unidades': sugestao_unidades,
            'sugestao_caixas': caixas_necessarias,
            'estoque_suficiente': False,
            'dias_cobertura_atual': estoque_atual / venda_media_dia if venda_media_dia > 0 else 0,
            'dias_cobertura_apos_pedido': (estoque_atual + sugestao_unidades) / venda_media_dia if venda_media_dia > 0 else float('inf'),
            'estrategia': estrategia,
            'observacao': observacao,
            'ponto_pedido': ponto_pedido,
            'estoque_ideal': estoque_ideal,
            'tendencia': tendencia,
            'motivo': motivo
        }
    
    def _analisar_tendencia(
        self,
        venda_7dias: int,
        venda_14dias: int,
        venda_30dias: int,
        venda_60dias: int
    ) -> Dict:
        """
        Analisa a tendência de vendas
        
        Returns:
            Dicionário com tipo de tendência e percentual
        """
        if venda_7dias == 0 or venda_14dias == 0:
            return {'tipo': 'estavel', 'percentual': 0, 'descricao': 'Sem dados suficientes'}
        
        # Calcula média diária de cada período
        media_7dias = venda_7dias / 7
        media_14dias = venda_14dias / 14
        media_30dias = venda_30dias / 30 if venda_30dias > 0 else media_14dias
        
        # Compara período mais recente (7 dias) com período anterior (14 dias)
        if media_14dias > 0:
            variacao = ((media_7dias - (venda_14dias - venda_7dias) / 7) / ((venda_14dias - venda_7dias) / 7)) * 100
        else:
            variacao = 0
        
        # Classifica tendência
        if variacao > 20:
            tipo = 'crescimento_forte'
            descricao = f'Crescimento forte ({variacao:.1f}%)'
        elif variacao > 5:
            tipo = 'crescimento'
            descricao = f'Crescimento moderado ({variacao:.1f}%)'
        elif variacao < -20:
            tipo = 'queda_forte'
            descricao = f'Queda forte ({variacao:.1f}%)'
        elif variacao < -5:
            tipo = 'queda'
            descricao = f'Queda moderada ({variacao:.1f}%)'
        else:
            tipo = 'estavel'
            descricao = f'Vendas estáveis ({variacao:.1f}%)'
        
        return {
            'tipo': tipo,
            'percentual': round(variacao, 2),
            'descricao': descricao,
            'media_7dias': round(media_7dias, 2),
            'media_14dias': round(media_14dias, 2),
            'media_30dias': round(media_30dias, 2)
        }
    
    def processar_arquivo(
        self,
        arquivo_entrada: str = "data/gerado.xlsx",
        arquivo_saida: str = "data/gerado_com_sugestao.xlsx"
    ) -> pd.DataFrame:
        """
        Processa o arquivo Excel e preenche a coluna de sugestão
        
        Args:
            arquivo_entrada: Caminho do arquivo de entrada
            arquivo_saida: Caminho do arquivo de saída
            
        Returns:
            DataFrame processado
        """
        
        print(f"📂 Lendo arquivo: {arquivo_entrada}")
        df = pd.read_excel(arquivo_entrada)
        
        print(f"📊 Total de linhas: {len(df)}")
        print(f"   Colunas: {list(df.columns)}\n")
        
        # Processa cada linha
        resultados = []
        
        for idx, row in df.iterrows():
            # Obtém ponto_pedido e estoque_ideal (pode ser NaN)
            ponto_pedido = int(row['ponto_pedido']) if pd.notna(row['ponto_pedido']) else None
            estoque_ideal = int(row['estoque_ideal']) if pd.notna(row['estoque_ideal']) else None
            
            resultado = self.calcular_sugestao_pedido(
                estoque_atual=int(row['estoque_atual']),
                venda_media_dia=float(row['venda_media_dia']),
                embalagem=int(row['embalagem']),
                venda_7dias=int(row['venda_acumulada_7dias']),
                venda_14dias=int(row['venda_acumulada_14dias']),
                venda_30dias=int(row['venda_acumulada_30dias']),
                venda_60dias=int(row['venda_acumulada_60dias']),
                ponto_pedido=ponto_pedido,
                estoque_ideal=estoque_ideal
            )
            
            resultados.append(resultado)
            
            # Log detalhado
            print(f"[{idx+1}/{len(df)}] Produto {row['codigo_interno']} - Loja {row['loja']}")
            print(f"  Estoque atual: {row['estoque_atual']} un")
            print(f"  Venda média/dia: {row['venda_media_dia']:.2f} un")
            print(f"  Cobertura atual: {resultado['dias_cobertura_atual']:.1f} dias")
            print(f"  Tendência: {resultado.get('tendencia', {}).get('descricao', 'N/A')}")
            print(f"  ➜ Sugestão: {resultado['sugestao_caixas']} caixas ({resultado['sugestao_unidades']} unidades)")
            print(f"  Motivo: {resultado['motivo']}")
            print()
        
        # Preenche a coluna sugestao com as unidades sugeridas
        df['sugestao'] = [r['sugestao_unidades'] for r in resultados]
        
        # Adiciona colunas extras com detalhes
        df['sugestao_caixas'] = [r['sugestao_caixas'] for r in resultados]
        df['dias_cobertura_atual'] = [round(r['dias_cobertura_atual'], 1) for r in resultados]
        df['dias_cobertura_apos'] = [round(r.get('dias_cobertura_apos_pedido', 0), 1) for r in resultados]
        df['estrategia_usada'] = [r.get('estrategia', 'N/A') for r in resultados]
        df['tendencia'] = [r.get('tendencia', {}).get('descricao', 'N/A') for r in resultados]
        df['motivo_sugestao'] = [r['motivo'] for r in resultados]
        
        # Formata codigo_interno com 7 dígitos (string com zeros à esquerda)
        df['codigo_interno'] = df['codigo_interno'].apply(lambda x: str(int(x)).zfill(7))
        
        # Formata loja com 3 dígitos (string com zeros à esquerda)
        df['loja'] = df['loja'].apply(lambda x: str(int(x)).zfill(3))
        
        # Salva arquivo com formatação de texto
        print(f"💾 Salvando arquivo: {arquivo_saida}")
        df.to_excel(arquivo_saida, index=False)
        
        # Força formato texto nas colunas codigo_interno e loja usando openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(arquivo_saida)
            ws = wb.active
            
            # Identifica as colunas
            headers = [cell.value for cell in ws[1]]
            col_codigo = headers.index('codigo_interno') + 1
            col_loja = headers.index('loja') + 1
            
            # Aplica formato texto (@) nas colunas
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_codigo).number_format = '@'
                ws.cell(row=row, column=col_loja).number_format = '@'
            
            wb.save(arquivo_saida)
        except Exception as e:
            print(f"⚠️  Aviso: Não foi possível aplicar formato texto: {e}")
        
        print(f"[OK] Arquivo processado com sucesso!")
        
        # Relatório resumido
        print("\n" + "="*70)
        print("📊 RESUMO DO PROCESSAMENTO")
        print("="*70)
        print(f"Total de produtos: {len(df)}")
        print(f"Produtos com sugestão > 0: {len(df[df['sugestao'] > 0])}")
        print(f"Produtos com estoque suficiente: {len(df[df['sugestao'] == 0])}")
        print(f"\nTotal de caixas sugeridas: {df['sugestao_caixas'].sum():.0f}")
        print(f"Total de unidades sugeridas: {df['sugestao'].sum():.0f}")
        print(f"\nMédia de dias de cobertura atual: {df['dias_cobertura_atual'].mean():.1f} dias")
        print(f"Média de dias de cobertura após pedido: {df['dias_cobertura_apos'].mean():.1f} dias")
        print("="*70)
        
        return df
    
    def gerar_relatorio_detalhado(self, df: pd.DataFrame) -> str:
        """Gera relatório detalhado em texto"""
        
        relatorio = []
        relatorio.append("="*70)
        relatorio.append("RELATÓRIO DETALHADO - SUGESTÕES DE PEDIDO")
        relatorio.append("="*70)
        relatorio.append("")
        
        for idx, row in df.iterrows():
            relatorio.append(f"PRODUTO: {row['codigo_interno']} | LOJA: {row['loja']}")
            relatorio.append("-"*70)
            relatorio.append(f"Estoque atual: {row['estoque_atual']} unidades")
            relatorio.append(f"Embalagem: {row['embalagem']} unidades/caixa")
            relatorio.append(f"")
            relatorio.append(f"Vendas:")
            relatorio.append(f"  - Média diária: {row['venda_media_dia']:.2f} un/dia")
            relatorio.append(f"  - 7 dias: {row['venda_acumulada_7dias']} un")
            relatorio.append(f"  - 14 dias: {row['venda_acumulada_14dias']} un")
            relatorio.append(f"  - 30 dias: {row['venda_acumulada_30dias']} un")
            relatorio.append(f"  - 60 dias: {row['venda_acumulada_60dias']} un")
            relatorio.append(f"")
            relatorio.append(f"Análise:")
            relatorio.append(f"  - Cobertura atual: {row['dias_cobertura_atual']:.1f} dias")
            relatorio.append(f"  - Tendência: {row['tendencia']}")
            relatorio.append(f"")
            relatorio.append(f"SUGESTÃO DE PEDIDO:")
            relatorio.append(f"  ➜ {row['sugestao_caixas']:.0f} caixas ({row['sugestao']:.0f} unidades)")
            relatorio.append(f"  - Cobertura após pedido: {row['dias_cobertura_apos']:.1f} dias")
            relatorio.append(f"  - Motivo: {row['motivo_sugestao']}")
            relatorio.append("")
            relatorio.append("="*70)
            relatorio.append("")
        
        return "\n".join(relatorio)


def main():
    """Função principal para executar o calculador"""
    
    print("="*70)
    print("  CALCULADOR DE SUGESTÕES DE PEDIDO")
    print("="*70)
    print()
    
    # Parâmetros
    print("⚙️  Parâmetros:")
    print(f"   - Dias de cobertura: 4 dias")
    print(f"   - Margem de segurança: 20%")
    print(f"   - Pedidos em múltiplos de embalagem (caixas fechadas)")
    print()
    
    # Inicializa calculador
    calculador = CalculadorPedido(dias_cobertura=4, margem_seguranca=1.2)
    
    # Processa arquivo
    try:
        df = calculador.processar_arquivo(
            arquivo_entrada="data/gerado.xlsx",
            arquivo_saida="data/gerado_com_sugestao.xlsx"
        )
        
        # Gera relatório detalhado
        relatorio = calculador.gerar_relatorio_detalhado(df)
        
        # Salva relatório
        with open("data/relatorio_sugestoes.txt", "w", encoding="utf-8") as f:
            f.write(relatorio)
        
        print(f"\n📄 Relatório detalhado salvo em: data/relatorio_sugestoes.txt")
        
    except FileNotFoundError:
        print("[ERRO] Arquivo 'data/gerado.xlsx' não encontrado!")
    except Exception as e:
        print(f"[ERRO] {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
