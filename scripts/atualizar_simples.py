#!/usr/bin/env python
# -*- coding: utf-8 -*-
import pandas as pd
import sys
import io
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers

# Configurar stdout para UTF-8 no Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Adiciona diretório raiz ao path
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from src.calculador_pedido import CalculadorPedido

def main():
    """Função principal para atualização rápida"""
    print("Atualizando coluna sugestao...")
    
    # Usar resultado_abc.xlsx que foi processado pelo tratamento_abc.py
    df = pd.read_excel('data/resultado_abc.xlsx')
    calc = CalculadorPedido(4, 1.2)

    sugestoes = []
    estrategias = []
    
    for _, row in df.iterrows():
        # Obtém ponto_pedido e estoque_ideal (já estão no resultado_abc.xlsx)
        ponto_pedido = float(row['ponto_pedido']) if pd.notna(row['ponto_pedido']) else None
        estoque_ideal = float(row['estoque_ideal']) if pd.notna(row['estoque_ideal']) else None
        estoque_atual = float(row['estoque']) if pd.notna(row['estoque']) else 0
        embalagem = int(row['embalagem']) if pd.notna(row['embalagem']) and row['embalagem'] > 0 else 1
        
        # Usar quantidade_vendida como proxy para venda_media_dia
        quantidade_vendida = float(row['quantidade_vendida']) if pd.notna(row['quantidade_vendida']) else 0
        venda_media_dia = quantidade_vendida  # Dados do dia
        
        # Cálculo simplificado sem usar o CalculadorPedido (evita divisão por zero)
        # Baseado na estratégia de balanceamento: 4-6 dias de cobertura
        
        if ponto_pedido and estoque_ideal and venda_media_dia > 0:
            # Calcular dias de cobertura dos valores do comprador
            dias_cobertura = (estoque_ideal - ponto_pedido) / venda_media_dia
            
            if 4 <= dias_cobertura <= 6:
                # COMPRADOR: valores adequados
                if estoque_atual < ponto_pedido:
                    sugestao = estoque_ideal - estoque_atual
                else:
                    sugestao = 0
                estrategia = 'COMPRADOR'
            elif dias_cobertura > 6:
                # GIRO_OTIMIZADO: ajustar para baixo (6 dias)
                estoque_alvo = ponto_pedido + (venda_media_dia * 6)
                if estoque_atual < ponto_pedido:
                    sugestao = estoque_alvo - estoque_atual
                else:
                    sugestao = 0
                estrategia = 'GIRO_OTIMIZADO_BAIXO'
            else:
                # GIRO_OTIMIZADO: ajustar para cima (4 dias)
                estoque_alvo = ponto_pedido + (venda_media_dia * 4)
                if estoque_atual < ponto_pedido:
                    sugestao = estoque_alvo - estoque_atual
                else:
                    sugestao = 0
                estrategia = 'GIRO_OTIMIZADO_ALTO'
        else:
            # GIRO_SAUDAVEL: usar 4 dias de cobertura
            if venda_media_dia > 0:
                estoque_alvo = venda_media_dia * 4
                if estoque_atual < estoque_alvo:
                    sugestao = estoque_alvo - estoque_atual
                else:
                    sugestao = 0
            else:
                sugestao = 0
            estrategia = 'GIRO_SAUDAVEL'
        
        # Arredondar para múltiplo de embalagem
        if sugestao > 0:
            sugestao = ((sugestao + embalagem - 1) // embalagem) * embalagem
        
        sugestoes.append(int(sugestao))
        estrategias.append(estrategia)

    df['sugestao'] = sugestoes
    df['estrategia'] = estrategias
    
    # codigo_interno e loja já estão padronizados no resultado_abc.xlsx
    # Apenas garantir que são strings
    df['codigo_interno'] = df['codigo_interno'].astype(str)
    df['loja'] = df['loja'].astype(str)
    
    # Salva em nova planilha
    arquivo_saida = 'data/sugestao_ia.xlsx'
    df.to_excel(arquivo_saida, index=False)
    
    # Força formato texto nas colunas codigo_interno e loja usando openpyxl
    wb = load_workbook(arquivo_saida)
    ws = wb.active
    
    # Identifica as colunas
    headers = [cell.value for cell in ws[1]]
    col_codigo = headers.index('codigo_interno') + 1
    col_loja = headers.index('loja') + 1
    
    # Aplica formato texto (@) nas colunas
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=col_codigo).number_format = '@'
        ws.cell(row=row, column=col_loja).number_format = '@'
    
    wb.save(arquivo_saida)
    
    print(f"Atualizado! {len([s for s in sugestoes if s > 0])} produtos com sugestao > 0")
    print(f"Total de unidades sugeridas: {sum(sugestoes)}")
    print(f"Arquivo salvo: {arquivo_saida}")

if __name__ == "__main__":
    main()
